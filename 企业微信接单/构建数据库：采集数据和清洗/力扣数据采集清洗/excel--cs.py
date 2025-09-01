from openpyxl import load_workbook

# 加载工作蒲
wb = load_workbook(r'cs.xlsx')
# 使用工作表
ws = wb["Sheet2"]

# 字段插入
ws['A1'] = 'Devansh Sharma'
ws['A2'] = 'hello world'
# 行列插入
ws.cell(row=2, column=2).value = 5
# 保存工作蒲
wb.save(r'cs.xlsx')
# 关闭工作蒲
wb.close()
print("运行结束！")

