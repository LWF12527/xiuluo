# coding=utf-8
import random
import re

import numpy as np
import pandas as pd
from openpyxl import load_workbook

# # 先分步插入到cs表，然后读取cs表，全部插入question表
# # 清洗数据
#
# # 加载AI判断题表
# # AI_pd_question_list = []
# # AI_pd_answer_list = []
# # data_AI_pd = pd.read_excel("AI.xlsx", sheet_name="判断题", header=None)[0].to_list()
# # while np.nan in data_AI_pd:
# #     data_AI_pd.remove(np.nan)
# # print(data_AI_pd)
# # for i in data_AI_pd:
# #     AI_pd_question_list.append(re.split(r'[？. （）]', i)[2])
# #     AI_pd_answer_list.append(re.split(r'[？. （）]', i)[-2])
#
# # # 加载算法判断题表
# # sf_pd_question_list = []
# # sf_pd_answer_list = []
# # data_sf_pd = pd.read_excel("算法.xlsx", sheet_name="判断题", header=None)[0].to_list()
# # while np.nan in data_sf_pd:
# #     data_sf_pd.remove(np.nan)
# # print(data_sf_pd)
# # for i in data_sf_pd:
# #     print(re.split(r'[.： ]', i))
# #     sf_pd_question_list.append(re.split(r'[.： ]', i)[2])
# #     sf_pd_answer_list.append(re.split(r'[.： ]', i)[-1])
#
# # # 加载前端判断题表
# # qd_pd_question_list = []
# # qd_pd_answer_list = []
# # data_qd_pd = pd.read_excel("前端.xlsx", sheet_name="判断题", header=None)[0].to_list()
# # while np.nan in data_qd_pd:
# #     data_qd_pd.remove(np.nan)
# # print(data_qd_pd)
# # n = 0
# # for i in data_qd_pd:
# #     if n % 2 == 0:
# #         qd_pd_question_list.append(re.split(r'[.。]', i)[-1])
# #     else:
# #         if re.split(r'[.。]', i)[0] == '错':
# #             qd_pd_answer_list.append("错误")
# #         else:
# #             qd_pd_answer_list.append("正确")
# #     n = n + 1
#
# # # 加载后端判断题表
# # hd_pd_question_list = []
# # hd_pd_answer_list = []
# # data_hd_pd = pd.read_excel("后端.xlsx", sheet_name="判断题", header=None)[0].to_list()
# # while np.nan in data_hd_pd:
# #     data_hd_pd.remove(np.nan)
# # print(data_hd_pd)
# # n = 0
# # for i in data_hd_pd:
# #     if n % 2 == 0:
# #         hd_pd_question_list.append(re.split(r'[.。：]', i)[-1])
# #     else:
# #         hd_pd_answer_list.append(re.split(r'[.。：]', i)[-2])
# #     n = n + 1
#
#
# # # 加载AI选择题表
# # AI_xz_answer_list = []
# # AI_xz_A_list = []
# # AI_xz_B_list = []
# # AI_xz_C_list = []
# # AI_xz_D_list = []
# # AI_xz_question_list = []
# # data_AI_xz = pd.read_excel("AI.xlsx", sheet_name="选择题", header=None)[0].to_list()
# # while np.nan in data_AI_xz:
# #     data_AI_xz.remove(np.nan)
# # print(data_AI_xz)
# # n = 0
# # for i in data_AI_xz:
# #     if n % 3 == 0:
# #         AI_xz_question_list.append(re.split(r'[.]', i)[-1])
# #     elif n % 3 == 2:
# #         AI_xz_answer_list.append(re.split(r'[.：]', i)[-2])
# #     else:
# #         # print(re.split(r'[ ]{2}', i))  # 两个空格拆分
# #         AI_xz_A_list.append(re.split(r'[ ]{2}', i)[0])
# #         AI_xz_B_list.append(re.split(r'[ ]{2}', i)[1])
# #         AI_xz_C_list.append(re.split(r'[ ]{2}', i)[2])
# #         AI_xz_D_list.append(re.split(r'[ ]{2}', i)[3])
# #     n = n + 1
# # # print(AI_xz_question_list, AI_xz_A_list, AI_xz_B_list, AI_xz_answer_list, end='\n')
#
#
# 加载算法选择题表
sf_xz_answer_list = []
sf_xz_A_list = []
sf_xz_B_list = []
sf_xz_C_list = []
sf_xz_D_list = []
sf_xz_question_list = []
data_sf_xz = pd.read_excel("算法.xlsx", sheet_name="选择题", header=None)[0].to_list()
while np.nan in data_sf_xz:
    data_sf_xz.remove(np.nan)
print(data_sf_xz)
n = 0
for i in data_sf_xz:
    if n % 6 == 0:
        sf_xz_question_list.append(re.split(r'[.]', i)[-1])
    elif n % 6 == 5:
        sf_xz_answer_list.append(re.split(r'[： ]', i)[-1])
    elif n % 6 == 1:
        sf_xz_A_list.append(i)
    elif n % 6 == 2:
        sf_xz_B_list.append(i)
    elif n % 6 == 3:
        sf_xz_C_list.append(i)
    elif n % 6 == 4:
        sf_xz_D_list.append(i)
    n = n + 1
print(sf_xz_question_list,sf_xz_answer_list,sf_xz_C_list,sf_xz_B_list,end='\n')
#
#
# # # 加载前端选择题表
# # qd_xz_answer_list = []
# # qd_xz_A_list = []
# # qd_xz_B_list = []
# # qd_xz_C_list = []
# # qd_xz_D_list = []
# # qd_xz_question_list = []
# # data_qd_xz = pd.read_excel("前端.xlsx", sheet_name="选择题", header=None)[0].to_list()
# # while np.nan in data_qd_xz:
# #     data_qd_xz.remove(np.nan)
# # # print(data_qd_xz)
# # n = 0
# # for i in data_qd_xz:
# #     if n % 6 == 0:
# #         qd_xz_question_list.append(re.split(r'[.]', i)[-1])
# #     elif n % 6 == 5:
# #         qd_xz_answer_list.append(re.split(r'[：. ]', i)[1])
# #     elif n % 6 == 1:
# #         qd_xz_A_list.append(i)
# #     elif n % 6 == 2:
# #         qd_xz_B_list.append(i)
# #     elif n % 6 == 3:
# #         qd_xz_C_list.append(i)
# #     elif n % 6 == 4:
# #         qd_xz_D_list.append(i)
# #     n = n + 1
# # # print(qd_xz_question_list, qd_xz_answer_list, qd_xz_C_list, qd_xz_B_list, end='\n')
#
#
# # # 加载后端选择题表
# # hd_xz_answer_list = []
# # hd_xz_A_list = []
# # hd_xz_B_list = []
# # hd_xz_C_list = []
# # hd_xz_D_list = []
# # hd_xz_question_list = []
# # data_hd_xz = pd.read_excel("后端.xlsx", sheet_name="选择题", header=None)[0].to_list()
# # while np.nan in data_hd_xz:
# #     data_hd_xz.remove(np.nan)
# # print(data_hd_xz)
# # n = 0
# # for i in data_hd_xz:
# #     if n % 6 == 0:
# #         hd_xz_question_list.append(re.split(r'[.]', i)[-1])
# #     elif n % 6 == 5:
# #         hd_xz_answer_list.append(re.split(r'[:. ]', i)[-3])
# #     elif n % 6 == 1:
# #         hd_xz_A_list.append(i)
# #     elif n % 6 == 2:
# #         hd_xz_B_list.append(i)
# #     elif n % 6 == 3:
# #         hd_xz_C_list.append(i)
# #     elif n % 6 == 4:
# #         hd_xz_D_list.append(i)
# #     n = n + 1
# # print(hd_xz_question_list, hd_xz_answer_list, hd_xz_C_list, hd_xz_B_list, end='\n')
#
# #
# # # 加载前端多选题表
# # qd_dx_answer_list = []
# # qd_dx_A_list = []
# # qd_dx_B_list = []
# # qd_dx_C_list = []
# # qd_dx_D_list = []
# # qd_dx_question_list = []
# # data_qd_dx = pd.read_excel("前端.xlsx", sheet_name="多选题", header=None)[0].to_list()
# # while np.nan in data_qd_dx:
# #     data_qd_dx.remove(np.nan)
# # print(data_qd_dx)
# # n = 0
# # for i in data_qd_dx:
# #     if n % 6 == 0:
# #         qd_dx_question_list.append(re.split(r'[.]', i)[-1])
# #     elif n % 6 == 5:
# #         qd_dx_answer_list.append(re.split(r'[: ]', i)[-1])
# #     elif n % 6 == 1:
# #         qd_dx_A_list.append(i)
# #     elif n % 6 == 2:
# #         qd_dx_B_list.append(i)
# #     elif n % 6 == 3:
# #         qd_dx_C_list.append(i)
# #     elif n % 6 == 4:
# #         qd_dx_D_list.append(i)
# #     n = n + 1
# # print(qd_dx_question_list, qd_dx_answer_list, qd_dx_C_list, qd_dx_B_list, end='\n')
#
#
# # # 加载前端多选题表
# # sf_dx_answer_list = []
# # sf_dx_A_list = []
# # sf_dx_B_list = []
# # sf_dx_C_list = []
# # sf_dx_D_list = []
# # sf_dx_question_list = []
# # data_sf_dx = pd.read_excel("算法.xlsx", sheet_name="多选题", header=None)[0].to_list()
# # while np.nan in data_sf_dx:
# #     data_sf_dx.remove(np.nan)
# # print(data_sf_dx)
# # n = 0
# # for i in data_sf_dx:
# #     if n % 6 == 0:
# #         sf_dx_question_list.append(re.split(r'[.]', i)[-1])
# #     elif n % 6 == 5:
# #         sf_dx_answer_list.append(re.split(r'[： ]', i)[-1])
# #     elif n % 6 == 1:
# #         sf_dx_A_list.append(i)
# #     elif n % 6 == 2:
# #         sf_dx_B_list.append(i)
# #     elif n % 6 == 3:
# #         sf_dx_C_list.append(i)
# #     elif n % 6 == 4:
# #         sf_dx_D_list.append(i)
# #     n = n + 1
# # print(sf_dx_question_list, sf_dx_answer_list, sf_dx_C_list, sf_dx_B_list, end='\n')
#
#
# # # 加载AI多选题表
# # AI_dx_answer_list = []
# # AI_dx_A_list = []
# # AI_dx_B_list = []
# # AI_dx_C_list = []
# # AI_dx_D_list = []
# # AI_dx_question_list = []
# # data_AI_dx = pd.read_excel("AI.xlsx", sheet_name="多选题", header=None)[0].to_list()
# # while np.nan in data_AI_dx:
# #     data_AI_dx.remove(np.nan)
# # print(data_AI_dx)
# # n = 0
# # for i in data_AI_dx:
# #     if n % 6 == 0:
# #         AI_dx_question_list.append(re.split(r'[.]', i)[-1])
# #     elif n % 6 == 5:
# #         AI_dx_answer_list.append(re.split(r'[： ]', i)[-1])
# #     elif n % 6 == 1:
# #         AI_dx_A_list.append(i)
# #     elif n % 6 == 2:
# #         AI_dx_B_list.append(i)
# #     elif n % 6 == 3:
# #         AI_dx_C_list.append(i)
# #     elif n % 6 == 4:
# #         AI_dx_D_list.append(i)
# #     n = n + 1
# # print(AI_dx_question_list, AI_dx_answer_list, AI_dx_C_list, AI_dx_B_list, end='\n')
#
# # # 加载后端多选题表
# # hd_dx_answer_list = []
# # hd_dx_A_list = []
# # hd_dx_B_list = []
# # hd_dx_C_list = []
# # hd_dx_D_list = []
# # hd_dx_question_list = []
# # data_hd_dx = pd.read_excel("后端.xlsx", sheet_name="多选题", header=None)[0].to_list()
# # while np.nan in data_hd_dx:
# #     data_hd_dx.remove(np.nan)
# # # print(data_hd_dx)
# # n = 0
# # for i in data_hd_dx:
# #     if n % 3 == 0:
# #         hd_dx_question_list.append(re.split(r'[.]', i)[-1])
# #     elif n % 3 == 2:
# #         hd_dx_answer_list.append(re.split(r'[.：]', i)[-1])
# #     else:
# #         hd_dx_A_list.append(re.split(r'[ ]', i)[0]+re.split(r'[ ]', i)[1])
# #         hd_dx_B_list.append(re.split(r'[ ]', i)[2]+re.split(r'[ ]', i)[3])
# #         hd_dx_C_list.append(re.split(r'[ ]', i)[4]+re.split(r'[ ]', i)[5])
# #         hd_dx_D_list.append(re.split(r'[ ]', i)[6]+re.split(r'[ ]', i)[7])
# #     n = n + 1
# # print(hd_dx_question_list, hd_dx_A_list, hd_dx_B_list, hd_dx_answer_list, end='\n')
# #
#
#
# # # 加载前端简答题表
# # jd_answer_list = []
# # jd_question_list = []
# # data_jd = pd.read_excel("算法.xlsx", sheet_name="简答题", header=None)[0].to_list()
# # while np.nan in data_jd:
# #     data_jd.remove(np.nan)
# # n = 0
# # for i in data_jd:
# #     if n % 2 == 0:
# #         jd_question_list.append(re.split(r'[.]', i)[-1])
# #     else:
# #         jd_answer_list.append(i)
# #     n=n+1
# # print(jd_question_list,jd_answer_list)
#
# 加载工作蒲
wb = load_workbook(r'question.xlsx')
# 使用工作表
ws = wb["Sheet1"]
# 字段插入
# ws['A1'] = 'Devansh Sharma'
grade_list = ['难', '中', '易']
for i in range(42, 57):
    ws['A{}'.format(i)] = sf_xz_question_list[i - 42]  # 问题
    ws['B{}'.format(i)] = "选择题"
    ws['C{}'.format(i)] = sf_xz_answer_list[i - 42]  # 答案
    # ws['E{}'.format(i)] = hd_dx_A_list[i - 141]  # A
    # ws['F{}'.format(i)] = hd_dx_B_list[i - 141]  # B
    # ws['G{}'.format(i)] = hd_dx_C_list[i - 141]  # C
    # ws['H{}'.format(i)] = hd_dx_D_list[i - 141]  # D
    ws['K{}'.format(i)] = '2'  # 分值
    ws['L{}'.format(i)] = grade_list[random.randint(0, 2)]  # 难度
    ws['M{}'.format(i)] = "算法"  # 所属知识点
    ws['N{}'.format(i)] = "算法"  # 知识类
    ws['O{}'.format(i)] = "力扣等面试题"  # 出处
    ws['P{}'.format(i)] = "算法"  # 知识点
    ws['Q{}'.format(i)] = "算法"  # 知识关键点
#
# # 保存工作蒲
wb.save(r'question.xlsx')
# 关闭工作蒲
wb.close()
print("\n运行结束！")
